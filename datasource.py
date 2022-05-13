import re
import time
from copy import copy

import openpyxl
import pandas as pd
import requests
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment

import config as conf


class GitData:
    def __init__(self):
        print('in Git')
        self.json_data = {}

    def __get_test_suite_info(self):
        try:
            res = requests.get(
                url=conf.GIT_INFO['suite_url'], verify=False,
                auth=(conf.GIT_INFO['user'], conf.GIT_INFO['token']))

            self.json_data = res.json()
            return self.json_data
        except Exception as e:
            print('Error:%s', str(e))
        return None

    def get_suite_files(self, suite_names):
        # returns python files of suites based on display name
        if not self.json_data:
            self.__get_test_suite_info()
        suit_files = []
        for suite_name in suite_names:
            for value in self.json_data.values():
                if suite_name in value:
                    suite_file = value[suite_name].rsplit('/', 1)[1]
                    suit_files.append(suite_file)
                    break
        return suit_files


class JenkinsData(GitData):
    def __init__(self, excel_file):
        super().__init__()
        print('In Jenkins Data')
        self.builid_name = {}
        self.__basic_info = None
        self.excel_file = excel_file

    def get_url(self, base_url):
        return base_url.replace("$build_info$",
                                conf.JENKINS_INFO['pipeline'] + '/' +
                                str(conf.JENKINS_INFO['build_id']))

    def get_suite_result(self):
        try:
            url_result = self.get_url(conf.URLS["url_result"])
            res = requests.get(
                url=url_result, verify=False,
                auth=(conf.JENKINS_INFO['user'], conf.JENKINS_INFO['token']))

            json_data = res.json()
            return self._get_tc_details(json_data)
        except Exception as e:
            print(str(e))
        return None

    def get_suite_input_params(self):
        try:
            url_result = self.get_url(conf.URLS["url_home"])
            res = requests.get(
                url=url_result, verify=False,
                auth=(conf.JENKINS_INFO['user'], conf.JENKINS_INFO['token']))

            json_data = res.json()
            return self._basic_info(json_data)
        except Exception as e:
            print(str(e))
        return None

    def __get_jobs_in_html(self):
        try:
            res = requests.get(
                url=self.get_url(conf.URLS["jobs"]), verify=False,
                auth=(conf.JENKINS_INFO['user'], conf.JENKINS_INFO['token']))

            jobs_data = res.json()

            jobs = ''
            for job in jobs_data["jobs"]:
                if job['name'] in ["jenkins-configuration",
                                   "master_report_merger", "Report_merger",
                                   "taas-examples"]:
                    continue
                jobs += '<option value="{}">{}</option>\n'.format(job['name'],
                                                                  job['name'])

            relevent_job = '<select name="name_jobs" class="combo">\n' + jobs + '</select>'
            return relevent_job
        except Exception as e:
            print(str(e))
        return None

    def get_all_jobs_html(self):
        try:
            jobs = self.__get_jobs_in_html()
            action_fun = "jenkin_job_builds"
            if "SDS" in self.excel_file:
                action_fun = "jenkin_job_builds?type=sds"
            jobs = '<div class="cnt_header"><form action=' + action_fun + ' method = "post"> \
                <div class="job_combo"> \
                <lable> Select pipeline/job:</lable>' + jobs + \
                   '<button class="button" id="jobs" value="param"> Submit \
                       <i class="fa fa-caret-down"></i> \
                   </button> </form></div>'
            print(jobs)
            return jobs
        except Exception as e:
            print('Error:%s', str(e))

    def _basic_info(self, json_data):
        parameters = []
        res_dict = {}
        epoch_time = json_data['timestamp']
        stime = '\n' + time.strftime('%Y-%m-%d',
                                     time.localtime(epoch_time / 1000))
        for param in json_data['actions']:
            if 'parameters' in param.keys():
                parameters = param
                break
        fvt_suites = []
        bvt_suites = []
        for param in parameters['parameters']:
            if param['name'] == 'OCP_CONSOLE_URL':
                # rack = re.search(r'isf-(.*).rtp', param['value'])
                if 'rack' in param['value']:
                    index = param['value'].find('rack')
                    res_dict['header'] = param['value'][
                                         index:index + 5] + '_' + stime
                elif 'sds' in param['value']:
                    index = param['value'].find('sds')
                    res_dict['header'] = param['value'][
                                         index:index + 6] + '_' + stime
                continue
            if 'BVT' in param['name']:
                bvt_suites = param['value'].split('\n')
                continue
            if 'FVT' in param['name']:
                if not len(param['value']):
                    continue
                suite_lst = param['value'].split('\n')
                fvt_suites += suite_lst
        res_dict['BVT'] = self.get_suite_files(bvt_suites)
        res_dict['FVT'] = self.get_suite_files(fvt_suites)
        self.__basic_info = res_dict
        return res_dict

    def __get_builds_in_html(self, job_name):
        try:
            url_build = self.get_url(conf.URLS["builds"])
            url_build = url_build.replace('$job$', job_name)
            res = requests.get(
                url_build, verify=False,
                auth=(conf.JENKINS_INFO['user'], conf.JENKINS_INFO['token']))

            jobs_data = res.json()

            jobs = ''
            for build in jobs_data["builds"]:
                self.builid_name[build['displayName']] = build['number']
                jobs += '<option value="{}">{}</option>\n'.format(
                    build['displayName'], build['displayName'])

            relevent_job = '<select name="name_builds" class="combo">\n' + jobs + '</select>'
            return relevent_job
        except Exception as e:
            print(str(e))
        return None

    def get_all_builds_html(self, job_name):
        try:
            builds = self.__get_builds_in_html(job_name)
            action_fun = "update_result"
            if "SDS" in self.excel_file:
                action_fun = "sds_update_result"

            builds = '<div class="cnt_header"><form action=' + action_fun + ' method = "post"> \
                <div class="build_combo"> \
                <lable> Select build to update result:</lable>' + builds + \
                     '<button class="button" id="jobs" value="param"> Submit \
                         <i class="fa fa-caret-down"></i> \
                     </button> </form></div>'
            print(builds)
            return builds
        except Exception as e:
            print('Error:%s', str(e))

    def _get_tc_details(self, json_data):
        bvt_lst = []
        fvt_lst = []
        all_suite_data = {}
        basic_info = self.__basic_info
        for tc in json_data['suites'][0]['cases']:
            suite_type = 'BVT'
            if tc['skipped']:
                status = "SKIPPED"
            elif tc['status'] == 'FIXED':
                status = 'PASSED'
            else:
                status = tc['status']
            # tc_name = tc['name'].split('[', 1)[0]
            suite_name = tc['className'].rsplit('.', 2)[-2] + '.py'
            suite_data = {'suite': suite_name,
                          'mthd_name': tc['name'],
                          'status': status}
            if suite_name in basic_info['BVT']:
                bvt_lst.append(suite_data)
            elif suite_name in basic_info['FVT']:
                fvt_lst.append(suite_data)
            else:
                print('Test case {} is not in BVT/FVT'.format(tc['className']))
        all_suite_data[conf.EXCEL_INFO['sheet_name'][0]] = bvt_lst
        all_suite_data[conf.EXCEL_INFO['sheet_name'][1]] = fvt_lst
        return all_suite_data


class ExcelData(JenkinsData):

    def __init__(self, excel_path):
        super().__init__(excel_path)
        print('in ExcelData class')
        self.df = None
        self.wb = None
        self.ws = None
        self.excel_path = excel_path
        self.cell_style = {}

    def read_excel(self, excel_file="Test.xlsx"):
        try:
            self.df = pd.read_excel(
                excel_file)

            print(self.df)
        except Exception as e:
            print('exception found %s', str(e))

    def get_html_data(self):
        self.read_excel()
        return self.df.to_html(classes='table table-striped')

    def get_xls_data(self, sheet):
        if self.wb is None:
            try:
                self.wb = openpyxl.load_workbook(self.excel_path)
            except:
                print('File does not exist, creating new file')
                self.wb = openpyxl.Workbook()
            self.cell_style = {
                "PASSED": self.create_named_style("Passed", 'FF00B050'),
                "FIXED": self.create_named_style("FIXED", 'FF00B050'),
                "FAILED": self.create_named_style("Failed", 'FFFF0000'),
                "REGRESSION": self.create_named_style("Regression", 'FFFF0000'),
                "SKIPPED": self.create_named_style('Skipped', '00FFFF00'),
                "DEFAULT": self.create_named_style('Other', '00010000')
            }
        try:
            self.ws = self.wb[sheet]
        except:
            print('sheet does not exist, creating new %s', sheet)
            self.ws = self.wb.create_sheet(sheet)

    def __excel_to_html(self, sheet_name):
        try:
            from xlsx2html import xlsx2html
            html_file = 'html_code/' + self.excel_path.rsplit('.', 1)[
                0] + '_' + sheet_name + '.html'
            xlsx2html(self.excel_path, html_file, sheet=sheet_name)
            return html_file
        except Exception as e:
            print('Error:%s', str(e))

    def excel_to_html(self, sheet_name):
        html_file = 'html_code/' + self.excel_path.rsplit('.', 1)[
            0] + '_' + sheet_name + '.html'
        # html_file = self.__excel_to_html(sheet_name)
        content = open(html_file, 'r').read()
        content = content.replace('\n', '')
        return self.hide_tag(content)

    def hide_tag(self, content):
        res = re.search(r'<body>(.*)</body>', content)

        return ('<div class="cnt_table">' + res.group() + '</div>')

    def update_jenkins_data(self, job_name, build_id, new_column=True):

        try:
            # suites_data = {'BVT':[{'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[isf-application-operator-controller-manager]',
            #                  'status': 'PASSED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[isf-data-protection-operator-controller-manager]',
            #                  'status': 'PASSED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[isf-prereq-operator-controller-manager]',
            #                  'status': 'PASSED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[isf-proxy]',
            #                  'status': 'PASSED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[isf-serviceability-operator-controller-manager]',
            #                  'status': 'PASSED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[isf-ui-dep]',
            #                  'status': 'PASSED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[isf-ui-operator-controller-manager]',
            #                  'status': 'PASSED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[isf-velero-operator-controller-manager]',
            #                  'status': 'FAILED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[logcollector]',
            #                  'status': 'PASSED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[oadp-operator]',
            #                  'status': 'FAILED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[velero]',
            #                  'status': 'FAILED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[noobaa-operator]',
            #                  'status': 'PASSED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[ocs-operator]',
            #                  'status': 'PASSED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[rook-ceph-operator]',
            #                  'status': 'PASSED'},
            #                 {'suite': 'test_sds_operators_bvt.py',
            #                  'mthd_name': 'test_validate_operator_health[ocs-metrics-exporter]',
            #                  'status': 'PASSED'}]}
            col_index = 0
            conf.JENKINS_INFO['pipeline'] = job_name
            conf.JENKINS_INFO['build_id'] = self.builid_name[build_id]
            param_info = self.get_suite_input_params()
            suites_data = self.get_suite_result()

            bd = Side(style='thin', color="000000")

            for sheet in conf.EXCEL_INFO['sheet_name']:
                if suites_data.get(sheet) is None:
                    continue
                self.get_xls_data(sheet)
                if new_column:
                    col_index = self.add_new_header(param_info['header'])
                for suite in suites_data.get(sheet):
                    row_index = self.get_tc_method_index(suite['suite'],
                                                         suite['mthd_name'])
                    if row_index == 0:
                        # add new row
                        row_index = self.ws.max_row + 1
                        if row_index < conf.EXCEL_INFO['first_row']:
                            row_index = conf.EXCEL_INFO['first_row']
                        self.ws[row_index][
                            conf.EXCEL_INFO['SerialNo']].value = row_index - \
                                                                 conf.EXCEL_INFO[
                                                                     'first_row'] + 1
                        self.ws[row_index][
                            conf.EXCEL_INFO['SerialNo']].border = Border(
                            left=bd, top=bd, right=bd, bottom=bd)
                        self.ws[row_index][
                            conf.EXCEL_INFO['SerialNo']].alignment = Alignment(
                            horizontal='center')
                        self.ws[row_index][conf.EXCEL_INFO['suites']].value = \
                        suite['suite']
                        self.ws[row_index][
                            conf.EXCEL_INFO['suites']].border = Border(left=bd,
                                                                       top=bd,
                                                                       right=bd,
                                                                       bottom=bd)
                        self.ws[row_index][
                            conf.EXCEL_INFO['scenarios']].value = suite[
                            'mthd_name']
                        self.ws[row_index][
                            conf.EXCEL_INFO['scenarios']].border = Border(
                            left=bd, top=bd, right=bd, bottom=bd)

                    self.ws[row_index][col_index].value = suite['status']
                    self.ws[row_index][col_index].style = self.cell_style[
                        suite['status']]

                self.wb.save(self.excel_path)
                self.__excel_to_html(sheet)
                print('sheet %s updated', sheet)
            self.wb.close()
            self.wb = None
            return True
        except Exception as e:
            print('Error:%s', str(e))
        return False

    def get_column_values(self, v):
        col_lst = []
        row_cells = self.ws[conf.EXCEL_INFO['suites']]
        for cell in row_cells:
            col_lst.append(cell.value)
        return col_lst

    def get_tc_method_index(self, tc, method):
        for row in range(3, self.ws.max_row + 1):
            sheet_tc = self.ws[row][conf.EXCEL_INFO['suites']].value
            sheet_scenario = self.ws[row][conf.EXCEL_INFO['scenarios']].value
            if sheet_tc and sheet_scenario and tc in sheet_tc and method in sheet_scenario:
                return row
        return 0

    def add_new_header(self, header_value):
        col = 0
        max_col = self.ws.max_column
        if max_col < 3:
            max_col = 3
            self.__set_header_value('A1', 'S.No.')
            self.__set_header_value('B1', 'TC Suite')
            self.__set_header_value('C1', 'TC Scenario')

        for col in range(2, max_col):
            if header_value == self.ws[1][col].value:
                return col
            elif not self.ws[1][col].value:
                self.ws[1][col].value = header_value
                return col
        col += 1
        # self.ws.insert_cols(idx=col)
        excel_col = chr(65 + col) + '1'
        blank_col = chr(65 + col) + '2'
        self.ws[excel_col].value = header_value
        self.ws[excel_col].font = copy(self.ws['A1'].font)
        self.ws[excel_col].border = copy(self.ws['A1'].border)
        self.ws[excel_col].alignment = copy(self.ws['A1'].alignment)
        self.ws[blank_col].border = copy(self.ws['A1'].border)

        return col

    def create_named_style(self, name, colour="000000", only_border=False):
        status_cell = NamedStyle(name=name)
        bd = Side(style='thin', color="000000")
        status_cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        status_cell.font = Font(size=11, color=colour)
        if name not in self.wb.named_styles:
            self.wb.add_named_style(status_cell)
        return name

    def __set_header_value(self, col, header_string):
        self.ws[col].value = header_string
        bd = Side(style='thin', color="000000")
        self.ws[col].border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.ws[col].font = Font(bold=True)
